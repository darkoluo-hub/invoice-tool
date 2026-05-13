#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
NDIS Invoice Generator UI v11 - Tax zero visible + optional PDF merge

Windows UI version for generating invoices from an Excel data table and an Excel invoice template.

Requires:
    pip install openpyxl pywin32 pypdf

How to use:
    1. Double-click this file on Windows.
    2. Select your invoice template .xlsx file.
    3. Select your invoice data .xlsx file.
    4. Select an output folder.
    5. Optional: tick Also export PDF.
    6. Optional: tick Merge Participant Information PDF and select the participant information PDF.
    7. Click Generate Invoices.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime, date, time
from pathlib import Path
import re
import shutil
import threading
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import win32com.client as win32
except ImportError:
    win32 = None

try:
    from pypdf import PdfReader, PdfWriter
except ImportError:
    try:
        from PyPDF2 import PdfReader, PdfWriter
    except ImportError:
        PdfReader = None
        PdfWriter = None

DATA_SHEET = "InvoiceData"
SERVICE_START_ROW = 22
SERVICE_END_ROW = 26
MAX_SERVICE_ROWS = SERVICE_END_ROW - SERVICE_START_ROW + 1
BLACK = 0
GREY = 8421504


# -----------------------------
# Invoice generation logic
# -----------------------------

def clean_filename(name: str) -> str:
    name = str(name or "").strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    if not name:
        name = "Invoice.xlsx"
    return name if name.lower().endswith(".xlsx") else f"{name}.xlsx"


def get_available_output_path(path: Path) -> Path:
    """Return a usable output path.

    If the target file is already open in Excel or locked by OneDrive, Excel may
    report "Cannot access ...". In that case we keep the original name if
    possible, otherwise create a safe numbered filename.
    """
    path = path.resolve()
    if not path.exists():
        return path

    try:
        path.unlink()
        return path
    except PermissionError:
        pass
    except OSError:
        pass

    stem = path.stem
    suffix = path.suffix or ".xlsx"
    for i in range(2, 1000):
        candidate = path.with_name(f"{stem}_{i}{suffix}")
        if not candidate.exists():
            return candidate
        try:
            candidate.unlink()
            return candidate
        except Exception:
            continue

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{stem}_{timestamp}{suffix}")


def format_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def format_time(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%I:%M %p").lstrip("0")
    if isinstance(value, time):
        return value.strftime("%I:%M %p").lstrip("0")
    return str(value).strip()


def group_digits_every_3(value) -> str:
    if value is None:
        return ""
    raw = str(value).strip()
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return raw
    return " ".join(digits[i:i + 3] for i in range(0, len(digits), 3))


def get_rows_from_data_file(data_path: Path) -> list[dict]:
    if load_workbook is None:
        raise RuntimeError("Missing openpyxl. Please install it with: pip install openpyxl")

    wb = load_workbook(data_path, data_only=True)
    if DATA_SHEET in wb.sheetnames:
        ws = wb[DATA_SHEET]
    else:
        ws = wb.active

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None or str(value).strip() == "" for value in row):
            continue
        record = {header: value for header, value in zip(headers, row) if header}
        if record.get("invoice_no"):
            rows.append(record)
    return rows


def build_description(item: dict) -> str:
    """Build the B-column service description.

    Default format:
        Worker @ DD/MM/YYYY HH:MM AM - HH:MM AM [Category] [Item]

    Optional columns supported:
        full_description / line_description / b_description
            If filled, this exact text is used in column B.

        line_detail / travel_detail / km_detail
            Text placed after the date and before the [Category] [Item].
            Good for km rows, e.g.
            "9.9 km calculated using Google Maps, $0.99 per km"

        line_suffix / description_suffix
            Text placed after [Category] [Item].
            Good for rows such as "Provider Travel".

    This supports examples like:
        Lina @ 23/04/2026 04:00 PM - 4:45 PM [Access...] [04_104_0125_6_1]
        Lina @ 23/04/2026 [Access...] [04_104_0125_6_1] Provider Travel
        Lina @ 23/04/2026 9.9 km calculated using Google Maps, $0.99 per km [Activity Based Transport...] [04_590_0125_6_1]
    """
    # If the user wants full manual control of the B-column text, use it exactly.
    full_description = (
        str(item.get("full_description") or "").strip()
        or str(item.get("line_description") or "").strip()
        or str(item.get("b_description") or "").strip()
    )
    if full_description:
        return full_description

    worker_name = (
        str(item.get("worker_name") or "").strip()
        or str(item.get("support_worker") or "").strip()
        or str(item.get("worker") or "").strip()
        or str(item.get("staff_name") or "").strip()
        or str(item.get("carer_name") or "").strip()
    )
    description = str(item.get("description") or "").strip()
    service_date = format_date(item.get("service_date"))
    start_time = format_time(item.get("start_time"))
    end_time = format_time(item.get("end_time"))
    support_category = str(item.get("support_category") or "").strip()
    support_item_number = str(item.get("support_item_number") or "").strip()
    line_detail = (
        str(item.get("line_detail") or "").strip()
        or str(item.get("travel_detail") or "").strip()
        or str(item.get("km_detail") or "").strip()
    )
    line_suffix = (
        str(item.get("line_suffix") or "").strip()
        or str(item.get("description_suffix") or "").strip()
    )

    name_text = worker_name or description

    if service_date:
        main_text = f"{name_text} @ {service_date}" if name_text else service_date
    else:
        main_text = name_text

    if line_detail:
        main_text = f"{main_text} {line_detail}".strip()
    elif start_time and end_time:
        main_text = f"{main_text} {start_time} - {end_time}".strip()

    parts = []
    if main_text:
        parts.append(main_text)
    if support_category:
        parts.append(f"[{support_category}]")
    if support_item_number:
        parts.append(f"[{support_item_number}]")
    if line_suffix:
        parts.append(line_suffix)
    return " ".join(parts).strip()


def clear_changeable_cells(ws) -> None:
    for row in range(9, 15):
        ws.Range(f"F{row}").Value = ""
    for row in range(7, 11):
        ws.Range(f"J{row}").Value = ""
    for row in range(SERVICE_START_ROW, SERVICE_END_ROW + 1):
        for col in ["B", "H", "I", "J", "K", "L"]:
            ws.Range(f"{col}{row}").Value = ""
    ws.Range("H28").Value = ""




def ensure_service_description_merges(ws) -> None:
    """Make sure each service description row is merged across B:F.

    Some template edits or Excel repairs can remove the B:F merged cells.
    We recreate them every time before writing B22:B26 so the generated
    invoice keeps the ShiftCare-style description layout.
    """
    for row_num in range(SERVICE_START_ROW, SERVICE_END_ROW + 1):
        rng = ws.Range(f"B{row_num}:F{row_num}")
        try:
            rng.UnMerge()
        except Exception:
            pass
        try:
            rng.Merge()
        except Exception:
            # If Excel says it is already merged, keep going.
            pass
        try:
            rng.WrapText = True
            rng.HorizontalAlignment = -4131  # xlLeft
            rng.VerticalAlignment = -4108    # xlCenter
        except Exception:
            pass


def ensure_tax_row_visible(ws) -> None:
    """Keep the tax total row visible even when tax is zero.

    Some templates use formulas or custom number formats that hide the tax row
    when the calculated tax is 0. For NDIS invoices, keeping H31/L31 visible is
    clearer for plan managers and easier for review.
    """
    try:
        ws.Rows(31).Hidden = False
        ws.Columns("H:H").Hidden = False
        ws.Columns("L:L").Hidden = False
    except Exception:
        pass

    try:
        if not str(ws.Range("H31").Value or "").strip():
            ws.Range("H31").Value = "Tax (10%)"
        ws.Range("H31").Font.Color = BLACK
    except Exception:
        pass

    try:
        # Force L31 to return 0 rather than blank when there is no tax.
        ws.Range("L31").Formula = "=SUM(K22:K26)"
        # Use a simple currency format with an explicit zero section.
        ws.Range("L31").NumberFormat = '[$$]#,##0.00;[Red]-[$$]#,##0.00;[$$]0.00'
        ws.Range("L31").Font.Color = BLACK
    except Exception:
        pass

def get_characters_font(cell, start: int, length: int):
    """
    Return the Font object for part of a cell text.

    Excel COM is a little inconsistent across Python/Excel versions.
    Some installations expose this as GetCharacters(), while others expose
    Characters(). We try the safest options first.
    """
    last_error = None

    attempts = [
        lambda: cell.GetCharacters(Start=start, Length=length).Font,
        lambda: cell.GetCharacters(start, length).Font,
        lambda: cell.Characters(Start=start, Length=length).Font,
        lambda: cell.Characters(start, length).Font,
    ]

    for attempt in attempts:
        try:
            return attempt()
        except Exception as exc:
            last_error = exc

    raise last_error


def set_partial_font(cell, start: int, length: int, *, color=None, bold=None) -> bool:
    if length <= 0:
        return True

    try:
        chars = get_characters_font(cell, start, length)
        if color is not None:
            chars.Color = color
        if bold is not None:
            chars.Bold = bool(bold)
        return True
    except Exception:
        # Do not stop invoice generation just because partial rich text fails.
        # The cell value is already written; the invoice remains usable.
        return False


def apply_header_rich_text(ws, invoice_no: str, issue_date: str, payment_date: str, ndis_number: str) -> None:
    f14_text = f"NDIA: {ndis_number}" if ndis_number else ""
    ws.Range("F14").Value = f14_text
    if ndis_number:
        set_partial_font(ws.Range("F14"), 1, len("NDIA: "), color=BLACK, bold=False)
        set_partial_font(ws.Range("F14"), len("NDIA: ") + 1, len(ndis_number), color=GREY, bold=False)

    j7_text = f"Tax Invoice #{invoice_no}"
    ws.Range("J7").Value = j7_text
    set_partial_font(ws.Range("J7"), 1, len("Tax Invoice"), color=BLACK, bold=True)
    invoice_part = f" #{invoice_no}"
    set_partial_font(ws.Range("J7"), len("Tax Invoice") + 1, len(invoice_part), color=GREY, bold=False)

    ws.Range("J8").Value = ""

    j9_label = "Issue Date: "
    ws.Range("J9").Value = f"{j9_label}{issue_date}"
    set_partial_font(ws.Range("J9"), 1, len(j9_label), color=BLACK, bold=False)
    set_partial_font(ws.Range("J9"), len(j9_label) + 1, len(issue_date), color=GREY, bold=False)

    j10_label = "Payment Date: "
    ws.Range("J10").Value = f"{j10_label}{payment_date}"
    set_partial_font(ws.Range("J10"), 1, len(j10_label), color=BLACK, bold=False)
    set_partial_font(ws.Range("J10"), len(j10_label) + 1, len(payment_date), color=GREY, bold=False)



def merge_pdf_with_participant_information(invoice_pdf_path: Path, participant_info_pdf_path: Path) -> Path:
    """Append the participant information PDF to the exported invoice PDF.

    The final PDF keeps the same invoice PDF filename. A temporary merged file is
    written first, then it replaces the original invoice PDF.
    """
    if PdfReader is None or PdfWriter is None:
        raise RuntimeError(
            "Missing PDF merge package. Please install it once with: pip install pypdf"
        )
    if not participant_info_pdf_path.exists():
        raise FileNotFoundError(f"Cannot find participant information PDF: {participant_info_pdf_path}")

    invoice_pdf_path = invoice_pdf_path.resolve()
    participant_info_pdf_path = participant_info_pdf_path.resolve()
    temp_path = invoice_pdf_path.with_name(f"{invoice_pdf_path.stem}__merged_tmp.pdf")

    writer = PdfWriter()
    for source_path in [invoice_pdf_path, participant_info_pdf_path]:
        reader = PdfReader(str(source_path))
        for page in reader.pages:
            writer.add_page(page)

    with open(temp_path, "wb") as f:
        writer.write(f)

    try:
        invoice_pdf_path.unlink()
    except FileNotFoundError:
        pass
    temp_path.replace(invoice_pdf_path)
    return invoice_pdf_path

def fill_invoice_with_excel(
    excel,
    template_path: Path,
    output_path: Path,
    invoice_rows: list[dict],
    export_pdf: bool = False,
    participant_info_pdf: Path | None = None,
) -> tuple[Path, Path | None]:
    first = invoice_rows[0]
    if len(invoice_rows) > MAX_SERVICE_ROWS:
        raise ValueError(
            f"Invoice {first.get('invoice_no')} has {len(invoice_rows)} service rows, "
            f"but the template only supports {MAX_SERVICE_ROWS}."
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path = get_available_output_path(output_path)
    shutil.copy2(template_path, output_path)

    wb = excel.Workbooks.Open(str(output_path.resolve()))
    try:
        ws = wb.Worksheets(1)
        clear_changeable_cells(ws)
        ensure_service_description_merges(ws)

        ws.Range("F9").Value = first.get("client_name") or ""
        ws.Range("F10").Value = first.get("client_address") or ""
        ws.Range("F11").Value = first.get("client_city") or ""
        ws.Range("F12").Value = first.get("client_phone") or ""
        ws.Range("F13").Value = first.get("client_email") or ""

        invoice_no = str(first.get("invoice_no") or "").strip()
        issue_date = format_date(first.get("issue_date"))
        payment_date = format_date(first.get("payment_date"))
        ndis_number = group_digits_every_3(first.get("ndis_number") or "")

        apply_header_rich_text(ws, invoice_no, issue_date, payment_date, ndis_number)
        ws.Range("H28").Value = f"Amount Due: {payment_date}"

        for idx, item in enumerate(invoice_rows):
            row_num = SERVICE_START_ROW + idx
            ws.Range(f"B{row_num}").Value = build_description(item)
            ws.Range(f"H{row_num}").Value = item.get("type") or "Hours"
            ws.Range(f"I{row_num}").Value = item.get("quantity") or 0
            ws.Range(f"J{row_num}").Value = item.get("unit_price") or 0
            ws.Range(f"K{row_num}").Value = item.get("tax") or 0

        for row_num in range(SERVICE_START_ROW, SERVICE_END_ROW + 1):
            ws.Range(f"L{row_num}").Formula = (
                f'=IF(COUNTA(B{row_num}:K{row_num})=0,"",J{row_num}*I{row_num})'
            )

        ensure_tax_row_visible(ws)

        wb.Application.CalculateFull()
        # The workbook is already a copied .xlsx file. Using Save() is safer
        # than SaveAs() here because SaveAs can fail when the output folder is
        # under OneDrive/Desktop or when the target filename already exists.
        wb.Save()

        pdf_path = None
        if export_pdf:
            pdf_path = get_available_output_path(output_path.with_suffix(".pdf"))
            # 0 = xlTypePDF. Export through Excel so the PDF keeps the same layout.
            wb.ExportAsFixedFormat(0, str(pdf_path.resolve()))
            if participant_info_pdf is not None:
                merge_pdf_with_participant_information(pdf_path, participant_info_pdf)
    finally:
        wb.Close(SaveChanges=False)

    return output_path, pdf_path


def generate_invoices(
    template_path: Path,
    data_path: Path,
    output_dir: Path,
    log_func=print,
    export_pdf: bool = False,
    participant_info_pdf: Path | None = None,
) -> int:
    if win32 is None:
        raise RuntimeError("Missing pywin32. Please install it with: pip install pywin32")
    if not template_path.exists():
        raise FileNotFoundError(f"Cannot find template file: {template_path}")
    if not data_path.exists():
        raise FileNotFoundError(f"Cannot find data file: {data_path}")
    if participant_info_pdf is not None and not participant_info_pdf.exists():
        raise FileNotFoundError(f"Cannot find participant information PDF: {participant_info_pdf}")

    rows = get_rows_from_data_file(data_path)
    grouped = defaultdict(list)
    for row in rows:
        invoice_no = str(row.get("invoice_no") or "").strip()
        grouped[invoice_no].append(row)

    if not grouped:
        raise ValueError("No invoice data found. Please check invoice_no column in your data file.")

    output_dir.mkdir(parents=True, exist_ok=True)
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        for invoice_no, invoice_rows in grouped.items():
            first = invoice_rows[0]
            output_filename = clean_filename(
                first.get("output_filename") or f"Invoice_{invoice_no}_{first.get('client_name', '')}.xlsx"
            )
            output_path = output_dir / output_filename
            actual_output_path, actual_pdf_path = fill_invoice_with_excel(
                excel,
                template_path,
                output_path,
                invoice_rows,
                export_pdf=export_pdf,
                participant_info_pdf=participant_info_pdf,
            )
            log_func(f"Generated Excel: {actual_output_path}")
            if actual_pdf_path:
                if participant_info_pdf is not None:
                    log_func(f"Generated merged PDF: {actual_pdf_path}")
                else:
                    log_func(f"Generated PDF: {actual_pdf_path}")
    finally:
        excel.Quit()

    return len(grouped)


# -----------------------------
# UI
# -----------------------------

class InvoiceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NDIS Invoice Generator")
        self.geometry("820x580")
        self.minsize(780, 540)

        self.template_var = tk.StringVar()
        self.data_var = tk.StringVar()
        self.output_var = tk.StringVar(value=str(Path.cwd() / "generated_invoices"))
        self.export_pdf_var = tk.BooleanVar(value=False)
        self.merge_participant_pdf_var = tk.BooleanVar(value=False)
        self.participant_pdf_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 12, "pady": 6}

        title = ttk.Label(self, text="NDIS Invoice Generator", font=("Segoe UI", 16, "bold"))
        title.pack(anchor="w", padx=14, pady=(14, 4))

        subtitle = ttk.Label(
            self,
            text="Select your template, invoice data file, and output folder. Then click Generate Invoices.",
            font=("Segoe UI", 10),
        )
        subtitle.pack(anchor="w", padx=14, pady=(0, 10))

        frame = ttk.Frame(self)
        frame.pack(fill="x", **pad)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="Invoice template (.xlsx)").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(frame, textvariable=self.template_var).grid(row=0, column=1, sticky="ew", pady=6)
        ttk.Button(frame, text="Browse", command=self.browse_template).grid(row=0, column=2, padx=(8, 0), pady=6)

        ttk.Label(frame, text="Invoice data (.xlsx)").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(frame, textvariable=self.data_var).grid(row=1, column=1, sticky="ew", pady=6)
        ttk.Button(frame, text="Browse", command=self.browse_data).grid(row=1, column=2, padx=(8, 0), pady=6)

        ttk.Label(frame, text="Output folder").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(frame, textvariable=self.output_var).grid(row=2, column=1, sticky="ew", pady=6)
        ttk.Button(frame, text="Browse", command=self.browse_output).grid(row=2, column=2, padx=(8, 0), pady=6)

        ttk.Checkbutton(
            frame,
            text="Also export PDF",
            variable=self.export_pdf_var,
        ).grid(row=3, column=1, sticky="w", pady=(2, 2))

        ttk.Checkbutton(
            frame,
            text="Merge Participant Information PDF into exported PDF",
            variable=self.merge_participant_pdf_var,
            command=self.on_merge_pdf_toggle,
        ).grid(row=4, column=1, sticky="w", pady=(2, 2))

        ttk.Label(frame, text="Participant information PDF").grid(row=5, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(frame, textvariable=self.participant_pdf_var).grid(row=5, column=1, sticky="ew", pady=6)
        ttk.Button(frame, text="Browse", command=self.browse_participant_pdf).grid(row=5, column=2, padx=(8, 0), pady=6)

        button_frame = ttk.Frame(self)
        button_frame.pack(fill="x", padx=12, pady=(4, 8))

        self.generate_button = ttk.Button(button_frame, text="Generate Invoices", command=self.start_generation)
        self.generate_button.pack(side="left")
        ttk.Button(button_frame, text="Open Output Folder", command=self.open_output_folder).pack(side="left", padx=(8, 0))
        ttk.Button(button_frame, text="Clear Log", command=self.clear_log).pack(side="left", padx=(8, 0))

        self.progress = ttk.Progressbar(self, mode="indeterminate")
        self.progress.pack(fill="x", padx=14, pady=(0, 8))

        ttk.Label(self, text="Log").pack(anchor="w", padx=14)
        log_frame = ttk.Frame(self)
        log_frame.pack(fill="both", expand=True, padx=14, pady=(4, 10))

        self.log_text = tk.Text(log_frame, height=12, wrap="word", font=("Consolas", 9))
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        status = ttk.Label(self, textvariable=self.status_var, relief="sunken", anchor="w")
        status.pack(fill="x", side="bottom")

        self.log("Ready. Please select your files and click Generate Invoices.")
        if load_workbook is None or win32 is None:
            self.log("Missing package detected. Install once: pip install openpyxl pywin32")
        if PdfReader is None or PdfWriter is None:
            self.log("PDF merge package not detected. Install once if you want PDF merging: pip install pypdf")

    def browse_template(self):
        path = filedialog.askopenfilename(
            title="Select invoice template",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.template_var.set(path)

    def browse_data(self):
        path = filedialog.askopenfilename(
            title="Select invoice data file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.data_var.set(path)

    def browse_output(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.output_var.set(path)

    def browse_participant_pdf(self):
        path = filedialog.askopenfilename(
            title="Select participant information PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if path:
            self.participant_pdf_var.set(path)
            self.merge_participant_pdf_var.set(True)
            self.export_pdf_var.set(True)

    def on_merge_pdf_toggle(self):
        if self.merge_participant_pdf_var.get():
            self.export_pdf_var.set(True)

    def open_output_folder(self):
        folder = Path(self.output_var.get()).expanduser()
        folder.mkdir(parents=True, exist_ok=True)
        try:
            import os
            os.startfile(str(folder))
        except Exception as exc:
            messagebox.showerror("Error", f"Cannot open folder:\n{exc}")

    def clear_log(self):
        self.log_text.delete("1.0", "end")

    def log(self, message: str):
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.update_idletasks()

    def validate_inputs(self) -> tuple[Path, Path, Path, Path | None] | None:
        template = Path(self.template_var.get().strip())
        data = Path(self.data_var.get().strip())
        output = Path(self.output_var.get().strip())
        participant_pdf = None
        if self.merge_participant_pdf_var.get():
            participant_pdf = Path(self.participant_pdf_var.get().strip())

        if not template.exists():
            messagebox.showwarning("Missing template", "Please select your invoice template .xlsx file.")
            return None
        if not data.exists():
            messagebox.showwarning("Missing data file", "Please select your invoice data .xlsx file.")
            return None
        if template.suffix.lower() != ".xlsx" or data.suffix.lower() != ".xlsx":
            messagebox.showwarning("Wrong file type", "Please select .xlsx files, not .xls or .csv.")
            return None
        if participant_pdf is not None:
            if not participant_pdf.exists():
                messagebox.showwarning("Missing participant PDF", "Please select the Participant Information PDF file.")
                return None
            if participant_pdf.suffix.lower() != ".pdf":
                messagebox.showwarning("Wrong PDF file type", "Please select a .pdf file for participant information.")
                return None
        return template, data, output, participant_pdf

    def start_generation(self):
        inputs = self.validate_inputs()
        if not inputs:
            return
        self.generate_button.configure(state="disabled")
        self.progress.start(10)
        self.status_var.set("Generating invoices...")
        self.log("Starting...")

        export_pdf = bool(self.export_pdf_var.get() or self.merge_participant_pdf_var.get())
        thread = threading.Thread(target=self._run_generation, args=(*inputs, export_pdf), daemon=True)
        thread.start()

    def _run_generation(
        self,
        template: Path,
        data: Path,
        output: Path,
        participant_pdf: Path | None,
        export_pdf: bool,
    ):
        try:
            count = generate_invoices(
                template,
                data,
                output,
                log_func=lambda m: self.after(0, self.log, m),
                export_pdf=export_pdf,
                participant_info_pdf=participant_pdf,
            )
            self.after(0, self._generation_done, count, output)
        except Exception as exc:
            details = traceback.format_exc()
            self.after(0, self._generation_failed, exc, details)

    def _generation_done(self, count: int, output: Path):
        self.progress.stop()
        self.generate_button.configure(state="normal")
        self.status_var.set(f"Done. {count} invoice file(s) generated.")
        self.log(f"Done. {count} invoice file(s) generated.")
        messagebox.showinfo("Done", f"{count} invoice file(s) generated.\n\nOutput folder:\n{output}")

    def _generation_failed(self, exc: Exception, details: str):
        self.progress.stop()
        self.generate_button.configure(state="normal")
        self.status_var.set("Failed")
        self.log("ERROR:")
        self.log(str(exc))
        self.log(details)
        messagebox.showerror("Error", f"Invoice generation failed:\n\n{exc}")


if __name__ == "__main__":
    app = InvoiceApp()
    app.mainloop()
